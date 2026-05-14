# -*- coding: utf-8 -*-
"""
Auditor IVA - AFIP/ARCA vs Libro IVA Sistema
Auditorías
Versión v5.3 multimes

Dos auditorías separadas:
1) Auditoría de Comprobantes: existencia documental por CUIT + Tipo + PV + Número.
2) Auditoría IVA del Mes: valida el IVA computado en el Libro IVA del período cargado,
   cruzándolo contra AFIP/ARCA como respaldo documental.

Incluye historial completo con descarga posterior de Excel/PDF.
"""
from __future__ import annotations

import base64
import json
import re
import unicodedata
from datetime import datetime, date
from io import BytesIO
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Tuple

import numpy as np
import pandas as pd
import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

try:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
    REPORTLAB_OK = True
except Exception:
    REPORTLAB_OK = False

APP_TITLE = "Auditor IVA"
APP_VERSION = "v5.3"
HIST_FILE = Path("historial_auditor_iva.json")
EXPORTS_DIR = Path("exports")
EXPORTS_DIR.mkdir(exist_ok=True)
TOLERANCIA_DEFAULT = 0.01

AUTH_USERS_DEFAULT: Dict[str, str] = {}
SOCIEDADES_CONFIG = {
    "La Forza Gastronómica SAS": "",
    "La Stazione Gastronómica SAS": "",
    "SIBI SA": "30-71142417-9",
}
SOCIEDADES_DEFAULT = list(SOCIEDADES_CONFIG.keys())

# =============================================================================
# Login
# =============================================================================

def _get_auth_users() -> Dict[str, str]:
    users = dict(AUTH_USERS_DEFAULT)
    try:
        if "APP_USER" in st.secrets and "APP_PASSWORD" in st.secrets:
            users[str(st.secrets["APP_USER"])] = str(st.secrets["APP_PASSWORD"])
        if "auth" in st.secrets:
            for k, v in dict(st.secrets["auth"]).items():
                users[str(k)] = str(v)
    except Exception:
        pass
    return users


def login_required() -> bool:
    if st.session_state.get("auth_ok"):
        with st.sidebar:
            st.success(f"Sesión iniciada:\n{st.session_state.get('auth_user', '')}")
            if st.button("Cerrar sesión"):
                st.session_state.pop("auth_ok", None)
                st.session_state.pop("auth_user", None)
                st.rerun()
        return True

    st.title(f"{APP_TITLE} · {APP_VERSION}")
    st.caption("Ingreso obligatorio")
    with st.form("login_form", clear_on_submit=False):
        user = st.text_input("Usuario")
        pwd = st.text_input("Contraseña", type="password")
        submit = st.form_submit_button("Ingresar", type="primary")
    users = _get_auth_users()
    if not users:
        st.error("No hay usuarios configurados. Definí APP_USER y APP_PASSWORD en secrets antes de usar la app.")
        return False
    if submit:
        if users.get(user) == pwd:
            st.session_state["auth_ok"] = True
            st.session_state["auth_user"] = user
            st.rerun()
        else:
            st.error("Usuario o contraseña incorrectos.")
    return False

# =============================================================================
# Normalización general
# =============================================================================

def remove_accents(s: str) -> str:
    return "".join(c for c in unicodedata.normalize("NFD", str(s)) if unicodedata.category(c) != "Mn")


def norm_text(x: Any) -> str:
    if pd.isna(x):
        return ""
    s = remove_accents(str(x)).upper().strip()
    s = re.sub(r"\s+", " ", s)
    return s


def norm_header(x: Any) -> str:
    s = norm_text(x)
    s = re.sub(r"[^A-Z0-9%]+", " ", s)
    return re.sub(r"\s+", " ", s).strip()


def parse_amount(x: Any) -> float:
    if pd.isna(x):
        return 0.0
    if isinstance(x, (int, float, np.integer, np.floating)) and not isinstance(x, bool):
        if np.isnan(x):
            return 0.0
        return float(x)
    s = str(x).strip().replace("\xa0", " ")
    s = s.replace("$", "").replace(" ", "")
    if s.lower() in {"", "nan", "none", "-"}:
        return 0.0
    neg = False
    if s.startswith("(") and s.endswith(")"):
        neg = True
        s = s[1:-1]
    if s.startswith("-"):
        neg = True
        s = s[1:]
    # Formatos argentinos y mixtos: 1.234.567,89 / 1234567.89
    if "," in s and "." in s:
        if s.rfind(",") > s.rfind("."):
            s = s.replace(".", "").replace(",", ".")
        else:
            s = s.replace(",", "")
    elif "," in s:
        s = s.replace(".", "").replace(",", ".")
    else:
        # si hay varios puntos, son miles
        if s.count(".") > 1:
            s = s.replace(".", "")
    try:
        val = float(s)
    except Exception:
        return 0.0
    return -val if neg else val


def parse_date(x: Any) -> Optional[pd.Timestamp]:
    if pd.isna(x):
        return None
    if isinstance(x, (pd.Timestamp, datetime, date)):
        try:
            return pd.to_datetime(x).normalize()
        except Exception:
            return None
    s = str(x).strip()
    if not s or s.lower() in {"nan", "none"}:
        return None
    # Evitar números de comprobante interpretados como fecha.
    if re.fullmatch(r"\d{8,}", s):
        return None
    dt = pd.to_datetime(s, dayfirst=True, errors="coerce")
    if pd.isna(dt):
        return None
    return pd.Timestamp(dt).normalize()


def norm_cuit(x: Any) -> str:
    s = re.sub(r"\D", "", str(x) if x is not None else "")
    if len(s) == 11:
        return s
    return s


def parse_numero(value: Any) -> Tuple[str, str]:
    """Devuelve (punto_venta, numero) normalizados desde un campo de número.
    Acepta 00008-000017552, 8-17552, 000017552, etc.
    """
    s = str(value if value is not None else "").strip()
    if s.lower() in {"", "nan", "none"}:
        return "", ""
    s = s.replace("/", "-").replace("–", "-").replace("—", "-")
    m = re.search(r"(\d+)\s*-\s*(\d+)", s)
    if m:
        return str(int(m.group(1))).zfill(5), str(int(m.group(2))).zfill(8)
    nums = re.findall(r"\d+", s)
    if len(nums) >= 2:
        return str(int(nums[-2])).zfill(5), str(int(nums[-1])).zfill(8)
    if len(nums) == 1:
        return "", str(int(nums[0])).zfill(8)
    return "", ""


def parse_pv_num(pv: Any, num: Any) -> Tuple[str, str]:
    pv_s = re.sub(r"\D", "", str(pv if pv is not None else ""))
    num_s = re.sub(r"\D", "", str(num if num is not None else ""))
    pv_norm = str(int(pv_s)).zfill(5) if pv_s else ""
    num_norm = str(int(num_s)).zfill(8) if num_s else ""
    return pv_norm, num_norm


def normalize_tipo(value: Any) -> str:
    s = norm_text(value)
    if not s:
        return ""
    # Si viene como "1 - Factura A".
    if " - " in s:
        s = s.split("-", 1)[1].strip()
    # Códigos directos del sistema.
    compact = re.sub(r"[^A-Z0-9]", "", s)
    if compact in {"FA", "FB", "FC", "FM", "FE"}:
        return compact
    if compact in {"NCA", "NCB", "NCC", "NCM", "NC"}:
        return compact
    if compact in {"NDA", "NDB", "NDC", "NDM", "ND"}:
        return compact
    if compact in {"TFA", "TFB", "TFC"}:
        return compact[-2:]
    # Textos AFIP.
    letter = ""
    m = re.search(r"\b([ABCM])\b", s)
    if m:
        letter = m.group(1)
    if "CREDITO" in s:
        return "NC" + letter if letter else "NC"
    if "DEBITO" in s:
        return "ND" + letter if letter else "ND"
    if "FACTURA" in s or "TIQUE" in s:
        return "F" + letter if letter else "F"
    return compact[:10]


def tipo_signo(tipo_norm: Any, tipo_original: Any = "") -> int:
    t = norm_text(str(tipo_norm) + " " + str(tipo_original))
    return -1 if ("NC" in t or "CREDITO" in t) else 1


def make_key(cuit: Any, tipo: Any, pv: Any, numero: Any) -> str:
    c = norm_cuit(cuit)
    t = normalize_tipo(tipo)
    p = re.sub(r"\D", "", str(pv or ""))
    n = re.sub(r"\D", "", str(numero or ""))
    p = str(int(p)).zfill(5) if p else ""
    n = str(int(n)).zfill(8) if n else ""
    if not c or not t or not n:
        return ""
    return f"{c}|{t}|{p}|{n}"


def is_valid_cuit(x: Any) -> bool:
    return len(norm_cuit(x)) == 11

# =============================================================================
# Lectura de archivos
# =============================================================================

def read_uploaded_file(uploaded) -> pd.DataFrame:
    """Lee CSV, XLSX o XLS y devuelve DataFrame raw sin limpiar.
    Para XLS viejo requiere xlrd en requirements.
    """
    name = getattr(uploaded, "name", "archivo")
    ext = Path(name).suffix.lower()
    data = uploaded.getvalue()
    bio = BytesIO(data)

    if ext == ".csv":
        text = None
        for enc in ["utf-8-sig", "utf-8", "latin-1", "cp1252"]:
            try:
                text = data.decode(enc)
                break
            except UnicodeDecodeError:
                continue
        if text is None:
            raise ValueError("No se pudo decodificar el CSV.")
        first = text.splitlines()[0] if text.splitlines() else ""
        sep = ";" if first.count(";") >= first.count(",") else ","
        return pd.read_csv(BytesIO(text.encode("utf-8")), sep=sep, header=None, engine="python")

    if ext == ".xlsx":
        return pd.read_excel(bio, header=None, engine="openpyxl")

    if ext == ".xls":
        try:
            return pd.read_excel(bio, header=None, engine="xlrd")
        except ImportError as e:
            raise ImportError(
                "Para leer .XLS antiguo hace falta xlrd. Verificá que requirements.txt incluya xlrd>=2.0.1."
            ) from e

    raise ValueError(f"Formato no soportado: {ext}. Subí CSV, XLSX o XLS.")


def find_header_row(raw: pd.DataFrame, required_terms: Iterable[str]) -> int:
    req = [norm_header(t) for t in required_terms]
    for i, row in raw.iterrows():
        joined = " | ".join(norm_header(v) for v in row.tolist())
        if all(t in joined for t in req):
            return int(i)
    raise ValueError(f"No se encontró fila de encabezado con términos: {required_terms}")


def dataframe_from_header(raw: pd.DataFrame, header_row: int) -> pd.DataFrame:
    df = raw.iloc[header_row + 1 :].copy()
    df.columns = raw.iloc[header_row].astype(str).tolist()
    df = df.dropna(how="all")
    return df


def col_find(df: pd.DataFrame, patterns: List[str], required: bool = True) -> Optional[str]:
    cols = list(df.columns)
    for pat in patterns:
        rx = re.compile(pat, re.I)
        for c in cols:
            if rx.search(norm_header(c)):
                return c
    if required:
        raise KeyError(f"No se encontró columna para patrones: {patterns}")
    return None

# =============================================================================
# Parsers AFIP / Libro IVA
# =============================================================================

def normalize_afip(raw: pd.DataFrame) -> pd.DataFrame:
    """Normaliza Mis Comprobantes Recibidos/Emitidos AFIP.
    Devuelve columnas estándar: fecha_comprobante, tipo, cuit, proveedor, pv, numero, iva, key.
    """
    # AFIP suele traer título en fila 0 y encabezado en fila 1.
    try:
        header_row = find_header_row(raw, ["Fecha", "Punto", "Numero", "Total IVA"])
    except Exception:
        header_row = 0
    df = dataframe_from_header(raw, header_row)

    fecha_col = col_find(df, [r"^FECHA$"])
    tipo_col = col_find(df, [r"^TIPO$"])
    pv_col = col_find(df, [r"PUNTO.*VENTA"])
    num_col = col_find(df, [r"NUMERO.*DESDE", r"NUMERO"])
    cuit_col = col_find(df, [r"NRO.*DOC.*EMISOR", r"CUIT.*EMISOR", r"CUIT.*PROVEEDOR"])
    prov_col = col_find(df, [r"DENOMINACION.*EMISOR", r"RAZON.*SOCIAL", r"PROVEEDOR"], required=False)
    iva_col = col_find(df, [r"TOTAL.*IVA", r"^IVA$", r"IVA REC"])
    total_col = col_find(df, [r"IMP.*TOTAL", r"TOTAL$"], required=False)

    out = pd.DataFrame()
    out["Fuente"] = "AFIP"
    out["Fecha_Comprobante"] = df[fecha_col].apply(parse_date)
    out["Tipo_Original"] = df[tipo_col].astype(str)
    out["Tipo"] = df[tipo_col].apply(normalize_tipo)
    out["CUIT"] = df[cuit_col].apply(norm_cuit)
    out["Proveedor"] = df[prov_col].astype(str).str.strip() if prov_col else ""
    pv_num = df.apply(lambda r: parse_pv_num(r[pv_col], r[num_col]), axis=1)
    out["PuntoVenta"] = [x[0] for x in pv_num]
    out["Numero"] = [x[1] for x in pv_num]
    raw_iva = df[iva_col].apply(parse_amount)
    out["IVA_Original"] = raw_iva
    out["Signo"] = out.apply(lambda r: tipo_signo(r["Tipo"], r["Tipo_Original"]), axis=1)
    out["IVA"] = out["Signo"] * raw_iva.abs()
    out["Importe_Total"] = df[total_col].apply(parse_amount) if total_col else 0.0
    out["Key"] = out.apply(lambda r: make_key(r["CUIT"], r["Tipo"], r["PuntoVenta"], r["Numero"]), axis=1)
    out = out[(out["CUIT"].apply(is_valid_cuit)) & (out["Numero"] != "") & (out["Key"] != "")].copy()
    return out.reset_index(drop=True)


def parse_libro_flexxus_layout(raw: pd.DataFrame) -> pd.DataFrame:
    """Parser específico para Libro IVA Compras del sistema/Flexxus observado:
    - Encabezado: PROVEEDOR, CAT. I.V.A., C.U.I.T., TIPO, NUMERO, ..., I.V.A.
    - La fecha del comprobante aparece como fila separada en columna A antes del proveedor.
    """
    header_row = find_header_row(raw, ["PROVEEDOR", "C U I T", "NUMERO", "I V A"])
    header = raw.iloc[header_row].tolist()
    data = raw.iloc[header_row + 1 :].copy().reset_index(drop=True)
    current_date: Optional[pd.Timestamp] = None
    rows: List[Dict[str, Any]] = []

    # Columnas por layout observado.
    prov_i, cuit_i, tipo_i, num_i, iva_i, total_i = 0, 5, 6, 7, 9, 16

    for _, r in data.iterrows():
        first = r.iloc[0] if len(r) > 0 else None
        dt = parse_date(first)
        cuit = r.iloc[cuit_i] if len(r) > cuit_i else None
        numero = r.iloc[num_i] if len(r) > num_i else None
        tipo = r.iloc[tipo_i] if len(r) > tipo_i else None
        iva = r.iloc[iva_i] if len(r) > iva_i else None

        if dt is not None and not is_valid_cuit(cuit):
            current_date = dt
            continue

        if not is_valid_cuit(cuit):
            continue
        pv, num = parse_numero(numero)
        tipo_norm = normalize_tipo(tipo)
        iva_raw = parse_amount(iva)
        sign = tipo_signo(tipo_norm, tipo)
        # Si el sistema ya lo trae negativo, abs conserva criterio de signo por tipo.
        iva_signed = sign * abs(iva_raw)
        rows.append({
            "Fuente": "Libro IVA",
            "Fecha_Comprobante": current_date,
            "Tipo_Original": str(tipo),
            "Tipo": tipo_norm,
            "CUIT": norm_cuit(cuit),
            "Proveedor": str(first).strip(),
            "PuntoVenta": pv,
            "Numero": num,
            "IVA_Original": iva_raw,
            "Signo": sign,
            "IVA": iva_signed,
            "Importe_Total": parse_amount(r.iloc[total_i]) if len(r) > total_i else 0.0,
        })

    out = pd.DataFrame(rows)
    if out.empty:
        raise ValueError("No se detectaron comprobantes válidos en el Libro IVA.")
    out["Key"] = out.apply(lambda r: make_key(r["CUIT"], r["Tipo"], r["PuntoVenta"], r["Numero"]), axis=1)
    out = out[(out["Key"] != "")].copy()
    return out.reset_index(drop=True)


def parse_libro_tabular(raw: pd.DataFrame) -> pd.DataFrame:
    """Parser genérico para CSV/XLSX tabular del sistema.
    Respeta la regla conocida: A proveedor, F CUIT, H número comprobante, J IVA,
    pero intenta usar encabezados cuando existen.
    """
    # Intentar detectar fila de encabezado; si no, usar posición.
    header_row = None
    try:
        header_row = find_header_row(raw, ["PROVEEDOR", "CUIT", "IVA"])
    except Exception:
        pass

    if header_row is not None:
        df = dataframe_from_header(raw, header_row)
        prov_col = col_find(df, [r"PROVEEDOR", r"RAZON", r"DENOMINACION"], required=False) or df.columns[0]
        cuit_col = col_find(df, [r"CUIT", r"C U I T"], required=False) or df.columns[5]
        tipo_col = col_find(df, [r"^TIPO$", r"COMPROBANTE"], required=False)
        num_col = col_find(df, [r"NUMERO", r"FACTURA", r"COMPROBANTE"], required=False) or df.columns[7]
        iva_col = col_find(df, [r"^I V A$", r"^IVA$", r"TOTAL.*IVA"], required=False) or df.columns[9]
        fecha_col = col_find(df, [r"FECHA"], required=False)
        total_col = col_find(df, [r"IMPORTE.*TOTAL", r"TOTAL$"], required=False)
    else:
        df = raw.copy().dropna(how="all")
        prov_col = df.columns[0]
        cuit_col = df.columns[5] if len(df.columns) > 5 else df.columns[0]
        tipo_col = df.columns[6] if len(df.columns) > 6 else None
        num_col = df.columns[7] if len(df.columns) > 7 else df.columns[0]
        iva_col = df.columns[9] if len(df.columns) > 9 else df.columns[-1]
        fecha_col = None
        total_col = df.columns[16] if len(df.columns) > 16 else None

    rows = []
    current_date = None
    for _, r in df.iterrows():
        if fecha_col is not None:
            current_date = parse_date(r.get(fecha_col))
        else:
            maybe_dt = parse_date(r.get(prov_col))
            if maybe_dt is not None and not is_valid_cuit(r.get(cuit_col)):
                current_date = maybe_dt
                continue

        cuit = norm_cuit(r.get(cuit_col))
        if not is_valid_cuit(cuit):
            continue
        tipo = r.get(tipo_col) if tipo_col is not None else ""
        pv, num = parse_numero(r.get(num_col))
        tipo_norm = normalize_tipo(tipo) or "F"  # fallback mínimo si el sistema no trae tipo
        iva_raw = parse_amount(r.get(iva_col))
        sign = tipo_signo(tipo_norm, tipo)
        rows.append({
            "Fuente": "Libro IVA",
            "Fecha_Comprobante": current_date,
            "Tipo_Original": str(tipo),
            "Tipo": tipo_norm,
            "CUIT": cuit,
            "Proveedor": str(r.get(prov_col, "")).strip(),
            "PuntoVenta": pv,
            "Numero": num,
            "IVA_Original": iva_raw,
            "Signo": sign,
            "IVA": sign * abs(iva_raw),
            "Importe_Total": parse_amount(r.get(total_col)) if total_col is not None else 0.0,
        })
    out = pd.DataFrame(rows)
    if out.empty:
        raise ValueError("No se detectaron comprobantes válidos en el Libro IVA.")
    out["Key"] = out.apply(lambda r: make_key(r["CUIT"], r["Tipo"], r["PuntoVenta"], r["Numero"]), axis=1)
    out = out[out["Key"] != ""].copy()
    return out.reset_index(drop=True)


def normalize_libro(raw: pd.DataFrame) -> pd.DataFrame:
    # Primero layout específico observado, luego genérico.
    try:
        return parse_libro_flexxus_layout(raw)
    except Exception:
        return parse_libro_tabular(raw)

def normalize_uploaded_files(files: List[Any], source_kind: str) -> pd.DataFrame:
    """Lee y normaliza uno o varios archivos de una misma fuente.

    source_kind:
    - "AFIP" para Mis Comprobantes.
    - "LIBRO" para Libro IVA del sistema.

    Devuelve un único DataFrame concatenado, conservando el nombre del archivo
    origen en la columna Archivo_Origen. Esto permite auditar varios meses en
    una sola corrida y mantener trazabilidad por archivo.
    """
    frames: List[pd.DataFrame] = []
    errors: List[str] = []
    for uploaded in files or []:
        name = getattr(uploaded, "name", "archivo_sin_nombre")
        try:
            raw = read_uploaded_file(uploaded)
            norm_df = normalize_afip(raw) if source_kind.upper() == "AFIP" else normalize_libro(raw)
            norm_df = norm_df.copy()
            norm_df["Archivo_Origen"] = name
            frames.append(norm_df)
        except Exception as exc:
            errors.append(f"{name}: {exc}")
    if errors:
        raise ValueError("No se pudieron procesar uno o más archivos:\n" + "\n".join(errors))
    if not frames:
        raise ValueError("No se cargaron archivos válidos para normalizar.")
    out = pd.concat(frames, ignore_index=True)
    if "Fecha_Comprobante" in out.columns:
        out["Periodo_Comprobante"] = pd.to_datetime(out["Fecha_Comprobante"], errors="coerce").dt.strftime("%m/%Y")
        out["Periodo_Comprobante"] = out["Periodo_Comprobante"].fillna("")
    return out


def file_names_label(files: List[Any]) -> str:
    names = [getattr(f, "name", "archivo_sin_nombre") for f in files or []]
    return ", ".join(names)


def run_auditoria_iva_consolidada(afip: pd.DataFrame, libro: pd.DataFrame, tolerancia: float) -> Dict[str, Any]:
    """Auditoría fiscal multiperíodo/consolidada.

    Se usa cuando el usuario carga varios meses juntos. No aplica filtro por
    mes/año, por lo que evita falsas alertas de "fuera de período" cuando el
    objetivo es auditar todos los archivos cargados como un único universo.
    """
    base = run_auditoria_comprobantes(afip, libro, tolerancia)
    detalle = base["detalle"].copy()
    metrics = dict(base["metrics"])
    metrics.update({
        "iva_computado_libro_mes": float(libro["IVA"].sum()),
        "iva_afip_encontrado_para_libro": float(detalle.loc[detalle["Estado"].isin(["OK", "IVA_DISTINTO"]), "IVA_AFIP"].sum()),
        "iva_libro_no_encontrado_afip": float(detalle.loc[detalle["Estado"] == "SOLO_LIBRO", "IVA_Libro"].sum()),
        "iva_afip_mes_no_registrado": float(detalle.loc[detalle["Estado"] == "SOLO_AFIP", "IVA_AFIP"].sum()),
        "comprobantes_fuera_periodo_libro": 0,
        "iva_fuera_periodo_libro": 0.0,
        "afip_mes_no_registrado_cantidad": int((detalle["Estado"] == "SOLO_AFIP").sum()),
        "modo_periodo": "Consolidado multiperíodo",
    })
    return {
        "tipo_auditoria": "Auditoría IVA del Mes",
        "metrics": metrics,
        "detalle": detalle,
        "resumen_estado": base["resumen_estado"],
        "resumen_proveedor": base["resumen_proveedor"],
        "duplicados": base["duplicados"],
        "afip_norm": afip,
        "libro_norm": libro,
        "afip_mes_no_registrado": detalle[detalle["Estado"] == "SOLO_AFIP"].copy(),
        "libro_fuera_periodo": pd.DataFrame(),
    }

# =============================================================================
# Auditorías
# =============================================================================

def aggregate_by_key(df: pd.DataFrame, fuente: str) -> Tuple[pd.DataFrame, pd.DataFrame]:
    dup = (
        df.groupby("Key")
        .agg(
            Cantidad=("Key", "size"),
            IVA=("IVA", "sum"),
            CUIT=("CUIT", "first"),
            Proveedor=("Proveedor", "first"),
            Tipo=("Tipo", "first"),
            PuntoVenta=("PuntoVenta", "first"),
            Numero=("Numero", "first"),
            Fecha_Comprobante=("Fecha_Comprobante", "first"),
        )
        .reset_index()
    )
    duplicates = dup[dup["Cantidad"] > 1].copy()
    duplicates["Fuente"] = fuente
    agg = dup.copy()
    return agg, duplicates


def run_auditoria_comprobantes(afip: pd.DataFrame, libro: pd.DataFrame, tolerancia: float) -> Dict[str, Any]:
    afip_agg, dup_afip = aggregate_by_key(afip, "AFIP")
    libro_agg, dup_libro = aggregate_by_key(libro, "Libro IVA")

    m = afip_agg.merge(libro_agg, on="Key", how="outer", suffixes=("_AFIP", "_Libro"), indicator=True)
    m["IVA_AFIP"] = m["IVA_AFIP"].fillna(0.0)
    m["IVA_Libro"] = m["IVA_Libro"].fillna(0.0)
    m["Diferencia"] = m["IVA_Libro"] - m["IVA_AFIP"]

    def estado(row):
        if row["_merge"] == "left_only":
            return "SOLO_AFIP"
        if row["_merge"] == "right_only":
            return "SOLO_LIBRO"
        if abs(row["Diferencia"]) > tolerancia:
            return "IVA_DISTINTO"
        return "OK"

    m["Estado"] = m.apply(estado, axis=1)
    for field in ["CUIT", "Proveedor", "Tipo", "PuntoVenta", "Numero", "Fecha_Comprobante"]:
        m[field] = m.get(f"{field}_AFIP").combine_first(m.get(f"{field}_Libro"))

    detail_cols = [
        "Estado", "CUIT", "Proveedor", "Tipo", "PuntoVenta", "Numero", "Fecha_Comprobante",
        "IVA_AFIP", "IVA_Libro", "Diferencia", "Key",
    ]
    detalle = m[detail_cols].copy()
    detalle = detalle.sort_values(["Estado", "CUIT", "PuntoVenta", "Numero"])

    resumen_estado = detalle.groupby("Estado").agg(
        Cantidad=("Estado", "size"),
        IVA_AFIP=("IVA_AFIP", "sum"),
        IVA_Libro=("IVA_Libro", "sum"),
        Diferencia=("Diferencia", "sum"),
        Diferencia_Abs=("Diferencia", lambda s: s.abs().sum()),
    ).reset_index()

    prov = detalle.groupby(["CUIT", "Proveedor"], dropna=False).agg(
        Cantidad=("Estado", "size"),
        OK=("Estado", lambda s: (s == "OK").sum()),
        Solo_AFIP=("Estado", lambda s: (s == "SOLO_AFIP").sum()),
        Solo_Libro=("Estado", lambda s: (s == "SOLO_LIBRO").sum()),
        IVA_Distinto=("Estado", lambda s: (s == "IVA_DISTINTO").sum()),
        IVA_AFIP=("IVA_AFIP", "sum"),
        IVA_Libro=("IVA_Libro", "sum"),
        Diferencia=("Diferencia", "sum"),
        Diferencia_Abs=("Diferencia", lambda s: s.abs().sum()),
    ).reset_index().sort_values("Diferencia_Abs", ascending=False)

    total_pos = detalle.loc[detalle["Diferencia"] > tolerancia, "Diferencia"].sum()
    total_neg = detalle.loc[detalle["Diferencia"] < -tolerancia, "Diferencia"].sum()
    total_abs = detalle.loc[detalle["Estado"] != "OK", "Diferencia"].abs().sum()
    metrics = {
        "iva_afip": float(detalle["IVA_AFIP"].sum()),
        "iva_libro": float(detalle["IVA_Libro"].sum()),
        "diferencia_neta": float(detalle["Diferencia"].sum()),
        "diferencias_positivas": float(total_pos),
        "diferencias_negativas": float(total_neg),
        "diferencia_bruta": float(total_abs),
        "cantidad_afip": int(len(afip)),
        "cantidad_libro": int(len(libro)),
        "claves_afip": int(len(afip_agg)),
        "claves_libro": int(len(libro_agg)),
        "claves_cruzadas": int(len(detalle)),
        "ok": int((detalle["Estado"] == "OK").sum()),
        "solo_afip": int((detalle["Estado"] == "SOLO_AFIP").sum()),
        "solo_libro": int((detalle["Estado"] == "SOLO_LIBRO").sum()),
        "iva_distinto": int((detalle["Estado"] == "IVA_DISTINTO").sum()),
        "observados": int((detalle["Estado"] != "OK").sum()),
        "duplicados_afip": int(len(dup_afip)),
        "duplicados_libro": int(len(dup_libro)),
        "iva_solo_afip": float(detalle.loc[detalle["Estado"] == "SOLO_AFIP", "IVA_AFIP"].sum()),
        "iva_solo_libro": float(detalle.loc[detalle["Estado"] == "SOLO_LIBRO", "IVA_Libro"].sum()),
        "iva_diferencia_mismo_comprobante": float(detalle.loc[detalle["Estado"] == "IVA_DISTINTO", "Diferencia"].sum()),
        "iva_diferencia_mismo_comprobante_abs": float(detalle.loc[detalle["Estado"] == "IVA_DISTINTO", "Diferencia"].abs().sum()),
    }
    return {
        "tipo_auditoria": "Auditoría de Comprobantes",
        "metrics": metrics,
        "detalle": detalle,
        "resumen_estado": resumen_estado,
        "resumen_proveedor": prov,
        "duplicados": pd.concat([dup_afip, dup_libro], ignore_index=True),
        "afip_norm": afip,
        "libro_norm": libro,
    }


def in_period(ts: Any, year: int, month: int) -> bool:
    dt = parse_date(ts)
    if dt is None:
        return False
    return int(dt.year) == int(year) and int(dt.month) == int(month)


def run_auditoria_iva_mes(afip: pd.DataFrame, libro: pd.DataFrame, year: int, month: int, tolerancia: float) -> Dict[str, Any]:
    """Universo principal: todo el Libro IVA cargado, asumido como reporte del período fiscal seleccionado.
    La fecha de cada línea se usa como fecha de comprobante para marcar comprobantes fuera del mes.
    AFIP se usa como respaldo documental por comprobante.
    """
    base = run_auditoria_comprobantes(afip, libro, tolerancia)
    detalle = base["detalle"].copy()

    # Marcas de período según fecha de comprobante del libro/afip si existe.
    detalle["Fecha_Comprobante"] = pd.to_datetime(detalle["Fecha_Comprobante"], errors="coerce")
    detalle["Comprobante_Fuera_Periodo"] = ~detalle["Fecha_Comprobante"].apply(lambda x: in_period(x, year, month))
    detalle["Comprobante_Fuera_Periodo"] = detalle["Comprobante_Fuera_Periodo"].fillna(True)

    # AFIP del mes no registrado: usando fecha comprobante AFIP.
    afip_mes = afip[afip["Fecha_Comprobante"].apply(lambda x: in_period(x, year, month))].copy()
    libro_keys = set(libro["Key"].astype(str))
    afip_mes_no_libro = afip_mes[~afip_mes["Key"].astype(str).isin(libro_keys)].copy()

    fuera_periodo_libro = libro[~libro["Fecha_Comprobante"].apply(lambda x: in_period(x, year, month))].copy()

    # Para IVA del mes, el total base es el Libro completo cargado.
    metrics = dict(base["metrics"])
    metrics.update({
        "iva_computado_libro_mes": float(libro["IVA"].sum()),
        "iva_afip_encontrado_para_libro": float(detalle.loc[detalle["Estado"].isin(["OK", "IVA_DISTINTO"]), "IVA_AFIP"].sum()),
        "iva_libro_no_encontrado_afip": float(detalle.loc[detalle["Estado"] == "SOLO_LIBRO", "IVA_Libro"].sum()),
        "iva_afip_mes_no_registrado": float(afip_mes_no_libro["IVA"].sum()),
        "comprobantes_fuera_periodo_libro": int(len(fuera_periodo_libro)),
        "iva_fuera_periodo_libro": float(fuera_periodo_libro["IVA"].sum()) if not fuera_periodo_libro.empty else 0.0,
        "afip_mes_no_registrado_cantidad": int(len(afip_mes_no_libro)),
    })
    return {
        "tipo_auditoria": "Auditoría IVA del Mes",
        "metrics": metrics,
        "detalle": detalle,
        "resumen_estado": base["resumen_estado"],
        "resumen_proveedor": base["resumen_proveedor"],
        "duplicados": base["duplicados"],
        "afip_norm": afip,
        "libro_norm": libro,
        "afip_mes_no_registrado": afip_mes_no_libro,
        "libro_fuera_periodo": fuera_periodo_libro,
    }


# =============================================================================
# Diagnóstico ejecutivo y plan de acción
# =============================================================================

MATERIALIDAD_REDONDEO_DEFAULT = 1.00


def _first_not_empty(series: pd.Series) -> str:
    for v in series.astype(str).tolist():
        if v and v.lower() not in {"nan", "none", ""}:
            return v
    return ""


def provider_display(cuit: Any, proveedor: Any) -> str:
    name = str(proveedor or "").strip()
    c = norm_cuit(cuit)
    return f"{name} ({c})" if c else name


def is_banco_nacion(row: pd.Series) -> bool:
    txt = norm_text(f"{row.get('Proveedor','')} {row.get('CUIT','')}")
    return "BANCO" in txt and ("NACION" in txt or "NAC ARGENTINA" in txt or "BNA" in txt)


def is_servicio_publico(row: pd.Series) -> bool:
    txt = norm_text(f"{row.get('Proveedor','')} {row.get('CUIT','')}")
    keys = [
        "EDEMSA", "EDEMSA", "COOP ELECTRICA", "COOPERATIVA ELECTRICA",
        "AGUA Y SANEAMIENTO", "AYSAM", "GAS", "ECOGAS", "TELECOM", "MUNICIPALIDAD",
    ]
    return any(k in txt for k in keys)


def categoria_solo_libro(row: pd.Series) -> str:
    if is_banco_nacion(row):
        return "Banco / operaciones recurrentes"
    if is_servicio_publico(row):
        return "Servicio público / posible desfase"
    fecha = parse_date(row.get("Fecha_Comprobante"))
    return "Verificar respaldo AFIP / comprobante" if fecha is not None else "Verificar fecha y respaldo"


def _fmt_pv_num(pv: Any, num: Any) -> str:
    p = str(pv or "").strip()
    n = str(num or "").strip()
    return f"{p}-{n}" if p else n


def enrich_detalle_for_report(detalle: pd.DataFrame) -> pd.DataFrame:
    if detalle is None or detalle.empty:
        return pd.DataFrame()
    df = detalle.copy()
    df["Fecha"] = pd.to_datetime(df.get("Fecha_Comprobante"), errors="coerce").dt.strftime("%d/%m/%Y")
    df["Fecha"] = df["Fecha"].fillna("")
    if "PuntoVenta" in df.columns and "Numero" in df.columns:
        df["Comprobante"] = df.apply(lambda r: _fmt_pv_num(r.get("PuntoVenta"), r.get("Numero")), axis=1)
    return df


def build_diagnostico_ejecutivo(result: Dict[str, Any], metadata: Dict[str, Any]) -> pd.DataFrame:
    m = result["metrics"]
    rows = []
    if result["tipo_auditoria"] == "Auditoría de Comprobantes":
        rows = [
            ["Auditoría ejecutada", "Comprobantes", "Control documental de existencia por CUIT + Tipo + Punto de Venta + Número."],
            ["Claves únicas AFIP / ARCA", m.get("claves_afip", 0), "Comprobantes normalizados detectados en la base externa."],
            ["Claves únicas Libro IVA", m.get("claves_libro", 0), "Comprobantes normalizados detectados en el sistema."],
            ["OK", m.get("ok", 0), "Existe en ambas fuentes y el IVA coincide dentro de tolerancia."],
            ["Sólo AFIP / ARCA", m.get("solo_afip", 0), "Comprobantes recibidos no encontrados en Libro. Posible crédito fiscal omitido."],
            ["Sólo Libro IVA", m.get("solo_libro", 0), "Comprobantes registrados en sistema no encontrados en AFIP. Revisar respaldo, carga o criterio de fecha."],
            ["IVA distinto", m.get("iva_distinto", 0), "Misma clave, distinto IVA. Separar redondeos de diferencias materiales."],
            ["Duplicados Libro", m.get("duplicados_libro", 0), "Riesgo de crédito fiscal duplicado si el mismo comprobante aparece más de una vez."],
            ["Diferencia bruta documental", m.get("diferencia_bruta", 0), "Riesgo bruto antes de compensar positivos y negativos."],
        ]
    else:
        rows = [
            ["Auditoría ejecutada", "IVA del Mes", "Valida el IVA computado en el Libro del período usando AFIP como respaldo documental."],
            ["IVA computado Libro del mes", m.get("iva_computado_libro_mes", 0), "Universo fiscal principal: el Libro IVA cargado para el período."],
            ["IVA AFIP respaldado para comprobantes del Libro", m.get("iva_afip_encontrado_para_libro", 0), "IVA respaldado por comprobantes encontrados en AFIP."],
            ["IVA Libro no encontrado en AFIP", m.get("iva_libro_no_encontrado_afip", 0), "Riesgo de crédito fiscal computado sin respaldo encontrado."],
            ["IVA AFIP del mes no registrado", m.get("iva_afip_mes_no_registrado", 0), "Potencial crédito fiscal omitido o pendiente de registración."],
            ["IVA con fecha fuera del período", m.get("iva_fuera_periodo_libro", 0), "Comprobantes registrados en el Libro del período pero con fecha visible distinta. Revisar fecha de recepción/registración y criterio fiscal."],
            ["Diferencia neta", m.get("diferencia_neta", 0), "Impacto final compensado. No reemplaza la auditoría documental."],
            ["Diferencia bruta no compensada", m.get("diferencia_bruta", 0), "Riesgo bruto real antes de compensaciones."],
        ]
    return pd.DataFrame(rows, columns=["Indicador", "Valor", "Interpretación contable"])


def build_claude_style_sheets(result: Dict[str, Any], metadata: Dict[str, Any], materialidad_redondeo: float = MATERIALIDAD_REDONDEO_DEFAULT) -> Dict[str, pd.DataFrame]:
    detalle = enrich_detalle_for_report(result.get("detalle", pd.DataFrame()))
    sheets: Dict[str, pd.DataFrame] = {}
    if detalle.empty:
        return sheets

    solo_afip = detalle[detalle["Estado"] == "SOLO_AFIP"].copy()
    solo_libro = detalle[detalle["Estado"] == "SOLO_LIBRO"].copy()
    iva_dist = detalle[detalle["Estado"] == "IVA_DISTINTO"].copy()

    if not solo_afip.empty:
        solo_afip["Diagnóstico / Acción"] = solo_afip.apply(
            lambda r: f"Verificar recepción y registrar en sistema si corresponde. IVA AFIP no cargado: {fmt_money(r.get('IVA_AFIP', 0))}", axis=1
        )
        sheets["SOLO_AFIP - No en Libro"] = solo_afip[["CUIT", "Proveedor", "Tipo", "PuntoVenta", "Numero", "Fecha", "IVA_AFIP", "Diagnóstico / Acción"]].sort_values("IVA_AFIP", ascending=False)

    if not solo_libro.empty:
        solo_libro["Categoría"] = solo_libro.apply(categoria_solo_libro, axis=1)
        solo_libro["Diagnóstico / Acción"] = solo_libro.apply(
            lambda r: "Controlar respaldo masivo/resumen y criterio de carga" if is_banco_nacion(r) else ("Controlar factura física/servicio y fecha de recepción" if is_servicio_publico(r) else "Verificar respaldo AFIP, CUIT, tipo, punto de venta, número y período"), axis=1
        )
        sheets["SOLO_LIBRO - Analisis"] = solo_libro[["CUIT", "Proveedor", "Tipo", "PuntoVenta", "Numero", "Fecha", "IVA_Libro", "Categoría", "Diagnóstico / Acción"]].sort_values("IVA_Libro", ascending=False)

    if not iva_dist.empty:
        iva_dist["Materialidad"] = np.where(iva_dist["Diferencia"].abs() <= materialidad_redondeo, "Redondeo / centavos", "Diferencia material")
        iva_dist["Acción"] = np.where(iva_dist["Materialidad"] == "Diferencia material", "Revisar alícuota/base imponible/carga manual", "Aceptar por tolerancia si política contable lo permite")
        sheets["IVA Distinto"] = iva_dist[["CUIT", "Proveedor", "Tipo", "PuntoVenta", "Numero", "Fecha", "IVA_AFIP", "IVA_Libro", "Diferencia", "Materialidad", "Acción"]].sort_values("Diferencia", key=lambda s: s.abs(), ascending=False)

    if result["tipo_auditoria"] == "Auditoría IVA del Mes":
        fuera = result.get("libro_fuera_periodo", pd.DataFrame()).copy()
        if fuera is not None and not fuera.empty:
            fuera["Fecha"] = pd.to_datetime(fuera.get("Fecha_Comprobante"), errors="coerce").dt.strftime("%d/%m/%Y").fillna("")
            fuera["Período fecha comprobante"] = pd.to_datetime(fuera.get("Fecha_Comprobante"), errors="coerce").dt.strftime("%m/%Y").fillna("")
            fuera["Acción"] = "Revisar fecha de recepción/registración y confirmar si corresponde computarlo en el período auditado. No rectificar automáticamente sin análisis contable."
            sheets["Fechas Fuera de Periodo"] = fuera[["CUIT", "Proveedor", "Tipo", "PuntoVenta", "Numero", "Fecha", "IVA", "Período fecha comprobante", "Acción"]].sort_values("IVA", ascending=False)

    return sheets


def build_plan_accion(result: Dict[str, Any], metadata: Dict[str, Any], materialidad_redondeo: float = MATERIALIDAD_REDONDEO_DEFAULT) -> pd.DataFrame:
    detalle = enrich_detalle_for_report(result.get("detalle", pd.DataFrame()))
    rows: List[List[Any]] = []
    idx = 1
    if detalle.empty:
        return pd.DataFrame(columns=["#", "Prioridad", "Acción", "Proveedor(es)", "IVA involucrado", "Responsable", "Estado"])

    # 1) Faltantes AFIP por proveedor: prioridad urgente por impacto.
    solo_afip = detalle[detalle["Estado"] == "SOLO_AFIP"].copy()
    if not solo_afip.empty:
        g = solo_afip.groupby(["CUIT", "Proveedor"], dropna=False).agg(
            Cantidad=("Estado", "size"), IVA=("IVA_AFIP", "sum"), MinFecha=("Fecha", _first_not_empty),
        ).reset_index().sort_values("IVA", ascending=False).head(10)
        for _, r in g.iterrows():
            prioridad = "URGENTE" if abs(r["IVA"]) >= 200000 else "MEDIA"
            rows.append([
                idx, prioridad,
                f"Verificar comprobantes recibidos en AFIP que no están en Libro. Si corresponden al período, registrar o justificar la no registración. Cantidad: {int(r['Cantidad'])}.",
                provider_display(r["CUIT"], r["Proveedor"]), float(r["IVA"]), "Pendiente", "Abierto"
            ])
            idx += 1

    # 2) Sólo Libro no encontrado, excluyendo Banco Nación informativo pero dejando servicios públicos.
    solo_libro = detalle[detalle["Estado"] == "SOLO_LIBRO"].copy()
    if not solo_libro.empty:
        solo_libro["EsBanco"] = solo_libro.apply(is_banco_nacion, axis=1)
        solo_libro["EsServicioPublico"] = solo_libro.apply(is_servicio_publico, axis=1)
        revisar = solo_libro[~solo_libro["EsBanco"]].copy()
        if not revisar.empty:
            g = revisar.groupby(["CUIT", "Proveedor"], dropna=False).agg(
                Cantidad=("Estado", "size"), IVA=("IVA_Libro", "sum"),
            ).reset_index().sort_values("IVA", ascending=False).head(8)
            for _, r in g.iterrows():
                prioridad = "MEDIA" if abs(r["IVA"]) >= 100000 else "BAJA"
                rows.append([
                    idx, prioridad,
                    f"Comprobantes registrados en Libro no encontrados en AFIP. Revisar respaldo, CUIT, tipo, punto de venta, número y período. Cantidad: {int(r['Cantidad'])}.",
                    provider_display(r["CUIT"], r["Proveedor"]), float(r["IVA"]), "Pendiente", "Abierto"
                ])
                idx += 1
        banco = solo_libro[solo_libro["EsBanco"]]
        if not banco.empty:
            rows.append([
                idx, "INFO",
                "Operaciones bancarias/recurrentes no encontradas individualmente en AFIP. Mantener control de respaldo masivo, resúmenes y criterio contable; no tratarlas como faltantes comerciales comunes.",
                "Banco Nación / operaciones bancarias", float(banco["IVA_Libro"].sum()), "Contabilidad", "Control periódico"
            ])
            idx += 1

    # 3) IVA distinto material.
    iva_dist = detalle[detalle["Estado"] == "IVA_DISTINTO"].copy()
    if not iva_dist.empty:
        mat = iva_dist[iva_dist["Diferencia"].abs() > materialidad_redondeo].sort_values("Diferencia", key=lambda s: s.abs(), ascending=False)
        if not mat.empty:
            for _, r in mat.head(5).iterrows():
                rows.append([
                    idx, "MEDIA",
                    f"Mismo comprobante con IVA distinto. Revisar base imponible, alícuota, exentos/percepciones o carga manual. Comp.: {_fmt_pv_num(r.get('PuntoVenta'), r.get('Numero'))}.",
                    provider_display(r.get("CUIT"), r.get("Proveedor")), float(r.get("Diferencia", 0)), "Pendiente", "Abierto"
                ])
                idx += 1
        redondeos = iva_dist[iva_dist["Diferencia"].abs() <= materialidad_redondeo]
        if not redondeos.empty:
            rows.append([
                idx, "BAJA",
                f"Separar {len(redondeos)} diferencias de centavos/redondeo. No mezclarlas con diferencias materiales; definir tolerancia de aceptación.",
                "Varios", float(redondeos["Diferencia"].abs().sum()), "Contabilidad", "Pendiente criterio"
            ])
            idx += 1

    # 4) Fechas fuera del período, sólo para IVA del Mes.
    if result["tipo_auditoria"] == "Auditoría IVA del Mes":
        fuera = result.get("libro_fuera_periodo", pd.DataFrame())
        if fuera is not None and not fuera.empty:
            rows.append([
                idx, "BAJA",
                "Comprobantes del Libro con fecha visible fuera del período. Revisar fecha de recepción/registración y confirmar criterio de cómputo. No definir rectificativa automática sin análisis del contador.",
                "Varios", float(fuera.get("IVA", pd.Series(dtype=float)).sum()), "Contabilidad", "Revisar criterio"
            ])
            idx += 1

    return pd.DataFrame(rows, columns=["#", "Prioridad", "Acción", "Proveedor(es)", "IVA involucrado", "Responsable", "Estado"])

# =============================================================================
# Exportaciones
# =============================================================================

def fmt_money(x: Any) -> str:
    try:
        return f"$ {float(x):,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
    except Exception:
        return "$ 0,00"


def auto_width(ws):
    for col in ws.columns:
        max_len = 0
        letter = get_column_letter(col[0].column)
        for cell in col:
            v = cell.value
            if v is not None:
                max_len = max(max_len, len(str(v)))
        ws.column_dimensions[letter].width = min(max(max_len + 2, 10), 48)


def write_df(ws, df: pd.DataFrame, start_row: int = 1, start_col: int = 1, title: Optional[str] = None):
    row = start_row
    if title:
        ws.cell(row=row, column=start_col, value=title).font = Font(bold=True, size=13)
        row += 2
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(border_style="thin", color="D9E2F3")
    for j, col in enumerate(df.columns, start_col):
        c = ws.cell(row=row, column=j, value=str(col))
        c.fill = header_fill
        c.font = header_font
        c.alignment = Alignment(horizontal="center")
        c.border = Border(top=thin, left=thin, right=thin, bottom=thin)
    for i, (_, record) in enumerate(df.iterrows(), row + 1):
        for j, col in enumerate(df.columns, start_col):
            val = record[col]
            if isinstance(val, pd.Timestamp):
                val = val.strftime("%d/%m/%Y")
            elif pd.isna(val):
                val = ""
            c = ws.cell(row=i, column=j, value=val)
            c.border = Border(top=thin, left=thin, right=thin, bottom=thin)
    auto_width(ws)


def export_excel(result: Dict[str, Any], metadata: Dict[str, Any]) -> bytes:
    """Exporta un informe más parecido a una auditoría profesional:
    Resumen Ejecutivo, hojas de observaciones por categoría y Plan de Acción.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Resumen Ejecutivo"

    title_fill = PatternFill("solid", fgColor="1F4E78")
    section_fill = PatternFill("solid", fgColor="D9EAF7")
    warning_fill = PatternFill("solid", fgColor="FCE4D6")
    ok_fill = PatternFill("solid", fgColor="E2F0D9")
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(border_style="thin", color="D9E2F3")

    ws.merge_cells("A1:D1")
    ws["A1"] = f"AUDITORÍA IVA — {metadata.get('sociedad','')} | PERÍODO {metadata.get('periodo','')}"
    ws["A1"].font = Font(bold=True, size=15, color="FFFFFF")
    ws["A1"].fill = title_fill
    ws["A1"].alignment = Alignment(horizontal="center")
    ws.merge_cells("A2:D2")
    ws["A2"] = f"CUIT: {metadata.get('cuit','')}  |  Usuario: {metadata.get('usuario','')}  |  Generado: {metadata.get('timestamp','')}"
    ws["A2"].font = Font(italic=True, color="666666")

    ws["A4"] = "INDICADORES CLAVE"
    ws["A4"].font = Font(bold=True, size=12)
    ws["A4"].fill = section_fill
    ws.merge_cells("A4:D4")

    diag = build_diagnostico_ejecutivo(result, metadata)
    write_df(ws, diag, start_row=5)

    # Diagnóstico corto por categoría.
    base_row = 5 + len(diag) + 3
    ws.cell(base_row, 1, "DIAGNÓSTICO CONTABLE").font = Font(bold=True, size=12)
    ws.cell(base_row, 1).fill = section_fill
    ws.merge_cells(start_row=base_row, start_column=1, end_row=base_row, end_column=4)
    m = result["metrics"]
    diagnostic_rows = [
        ["Diferencia neta", fmt_money(m.get("diferencia_neta", 0)), "Resultado compensado: no debe usarse como única conclusión."],
        ["Diferencia bruta", fmt_money(m.get("diferencia_bruta", 0)), "Mide el volumen real de diferencias antes de compensar."],
        ["Sólo AFIP", f"{m.get('solo_afip', 0)} comp. | {fmt_money(m.get('iva_solo_afip', 0))}", "Posibles comprobantes no registrados o crédito fiscal omitido."],
        ["Sólo Libro", f"{m.get('solo_libro', 0)} comp. | {fmt_money(m.get('iva_solo_libro', 0))}", "Revisar respaldo, datos de carga y criterio de período."],
        ["IVA distinto", f"{m.get('iva_distinto', 0)} comp. | {fmt_money(m.get('iva_diferencia_mismo_comprobante_abs', 0))}", "Separar diferencias materiales de redondeos."],
    ]
    write_df(ws, pd.DataFrame(diagnostic_rows, columns=["Categoría", "Valor", "Lectura / Riesgo"]), start_row=base_row + 1)

    # Hojas estilo informe externo.
    sheets = build_claude_style_sheets(result, metadata)
    sheets["Plan de Acción"] = build_plan_accion(result, metadata)

    # Hojas técnicas originales al final para trazabilidad completa.
    sheets.update({
        "Resumen por Estado": result.get("resumen_estado", pd.DataFrame()),
        "Diferencias por CUIT": result.get("resumen_proveedor", pd.DataFrame()),
        "Detalle Comprobantes": result.get("detalle", pd.DataFrame()),
        "Duplicados": result.get("duplicados", pd.DataFrame()),
        "AFIP Normalizado": result.get("afip_norm", pd.DataFrame()),
        "Libro Normalizado": result.get("libro_norm", pd.DataFrame()),
    })
    if result["tipo_auditoria"] == "Auditoría IVA del Mes":
        sheets["AFIP Mes no Registrado"] = result.get("afip_mes_no_registrado", pd.DataFrame())
        sheets["Libro Fuera Periodo"] = result.get("libro_fuera_periodo", pd.DataFrame())

    for name, df in sheets.items():
        wsx = wb.create_sheet(name[:31])
        if df is not None and not df.empty:
            title = None
            if name == "Plan de Acción":
                title = f"PLAN DE ACCIÓN — {metadata.get('sociedad','')} — {metadata.get('periodo','')}"
            write_df(wsx, df.reset_index(drop=True), title=title)
            # Formato especial para prioridad en Plan de Acción.
            if name == "Plan de Acción":
                wsx.freeze_panes = "A3"
                for row in range(3, wsx.max_row + 1):
                    val = str(wsx.cell(row, 2).value or "").upper()
                    if val == "URGENTE":
                        fill = PatternFill("solid", fgColor="F4CCCC")
                    elif val == "MEDIA":
                        fill = PatternFill("solid", fgColor="FFF2CC")
                    elif val == "BAJA":
                        fill = PatternFill("solid", fgColor="D9EAF7")
                    else:
                        fill = PatternFill("solid", fgColor="E2F0D9")
                    for col in range(1, min(wsx.max_column, 7) + 1):
                        wsx.cell(row, col).fill = fill
        else:
            wsx["A1"] = "Sin datos"
        auto_width(wsx)

    for sh in wb.worksheets:
        try:
            sh.freeze_panes = "A2" if sh.title != "Resumen Ejecutivo" else "A5"
        except Exception:
            pass
        for row in sh.iter_rows():
            for cell in row:
                if isinstance(cell.value, (int, float)) and ("IVA" in str(sh.title).upper() or cell.column >= 5):
                    cell.number_format = '#,##0.00'
                cell.alignment = Alignment(vertical="top", wrap_text=True)

    auto_width(ws)
    bio = BytesIO()
    wb.save(bio)
    return bio.getvalue()

def export_pdf(result: Dict[str, Any], metadata: Dict[str, Any]) -> Optional[bytes]:
    if not REPORTLAB_OK:
        return None
    bio = BytesIO()
    doc = SimpleDocTemplate(bio, pagesize=landscape(A4), rightMargin=1*cm, leftMargin=1*cm, topMargin=1*cm, bottomMargin=1*cm)
    styles = getSampleStyleSheet()
    story = []
    story.append(Paragraph("Auditor IVA", styles["Title"]))
    story.append(Paragraph(f"{result['tipo_auditoria']} · {metadata.get('sociedad','')} · Período {metadata.get('periodo','')}", styles["Normal"]))
    story.append(Paragraph(f"Usuario: {metadata.get('usuario','')} · Generado: {metadata.get('timestamp','')}", styles["Normal"]))
    story.append(Spacer(1, 0.3*cm))

    metrics = result["metrics"]
    if result["tipo_auditoria"] == "Auditoría de Comprobantes":
        data = [
            ["Indicador", "Valor"],
            ["Claves únicas AFIP", metrics.get("claves_afip", 0)],
            ["Claves únicas Libro", metrics.get("claves_libro", 0)],
            ["OK", metrics.get("ok", 0)],
            ["Observados", metrics.get("observados", 0)],
            ["Sólo AFIP", metrics.get("solo_afip", 0)],
            ["Sólo Libro", metrics.get("solo_libro", 0)],
            ["IVA distinto", metrics.get("iva_distinto", 0)],
            ["Duplicados Libro", metrics.get("duplicados_libro", 0)],
            ["IVA AFIP no cargado", fmt_money(metrics.get("iva_solo_afip", 0))],
            ["IVA Libro sin AFIP", fmt_money(metrics.get("iva_solo_libro", 0))],
            ["Diferencia bruta documental", fmt_money(metrics.get("diferencia_bruta", 0))],
        ]
    else:
        data = [
            ["Indicador", "Valor"],
            ["IVA computado Libro del mes", fmt_money(metrics.get("iva_computado_libro_mes", 0))],
            ["IVA AFIP respaldado", fmt_money(metrics.get("iva_afip_encontrado_para_libro", 0))],
            ["IVA Libro no encontrado AFIP", fmt_money(metrics.get("iva_libro_no_encontrado_afip", 0))],
            ["IVA AFIP del mes no registrado", fmt_money(metrics.get("iva_afip_mes_no_registrado", 0))],
            ["IVA fuera del período", fmt_money(metrics.get("iva_fuera_periodo_libro", 0))],
            ["Diferencia neta", fmt_money(metrics.get("diferencia_neta", 0))],
            ["Diferencia bruta no compensada", fmt_money(metrics.get("diferencia_bruta", 0))],
            ["OK", metrics.get("ok", 0)],
            ["Sólo AFIP", metrics.get("solo_afip", 0)],
            ["Sólo Libro", metrics.get("solo_libro", 0)],
            ["IVA distinto", metrics.get("iva_distinto", 0)],
        ]
    tbl = Table(data, colWidths=[10*cm, 7*cm])
    tbl.setStyle(TableStyle([
        ("BACKGROUND", (0,0), (-1,0), colors.HexColor("#1F4E78")),
        ("TEXTCOLOR", (0,0), (-1,0), colors.white),
        ("GRID", (0,0), (-1,-1), 0.25, colors.grey),
        ("FONTNAME", (0,0), (-1,0), "Helvetica-Bold"),
        ("ALIGN", (1,1), (1,-1), "RIGHT"),
    ]))
    story.append(tbl)
    story.append(Spacer(1, 0.4*cm))
    story.append(Paragraph("Nota: el PDF es ejecutivo. El Excel incluye Resumen Ejecutivo, observaciones por categoría y Plan de Acción.", styles["Italic"]))
    doc.build(story)
    return bio.getvalue()

# =============================================================================
# Historial
# =============================================================================

def load_history() -> List[Dict[str, Any]]:
    if HIST_FILE.exists():
        try:
            data = json.loads(HIST_FILE.read_text(encoding="utf-8"))
            return data if isinstance(data, list) else []
        except Exception:
            return []
    return []


def save_history(items: List[Dict[str, Any]]):
    HIST_FILE.write_text(json.dumps(items[:300], ensure_ascii=False, indent=2), encoding="utf-8")


def _bytes_to_b64(value: Optional[bytes]) -> str:
    if not value:
        return ""
    try:
        return base64.b64encode(value).decode("utf-8")
    except Exception:
        return ""


def _b64_to_bytes(value: Any) -> Optional[bytes]:
    if not value:
        return None
    try:
        return base64.b64decode(str(value).encode("utf-8"))
    except Exception:
        return None


def _read_file_if_exists(path_value: Any) -> Optional[bytes]:
    if not path_value:
        return None
    try:
        p = Path(str(path_value))
        if p.exists() and p.is_file():
            return p.read_bytes()
    except Exception:
        return None
    return None


def save_run_to_history(result: Dict[str, Any], metadata: Dict[str, Any], excel_bytes: bytes, pdf_bytes: Optional[bytes]) -> Dict[str, Any]:
    """Guarda la corrida completa.

    Además de guardar path local, embebe los bytes del Excel/PDF en base64. Esto evita
    que el historial quede con "archivo no disponible" cuando Streamlit Cloud reinicia
    y pierde archivos temporales de /exports.
    """
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    safe_tipo = "comprobantes" if "Comprobantes" in result["tipo_auditoria"] else "iva_mes"
    excel_name = f"auditoria_iva_{safe_tipo}_{metadata.get('periodo','').replace('/','_')}_{stamp}.xlsx"
    pdf_name = f"auditoria_iva_{safe_tipo}_{metadata.get('periodo','').replace('/','_')}_{stamp}.pdf"
    excel_path = EXPORTS_DIR / excel_name
    pdf_path = EXPORTS_DIR / pdf_name
    excel_path.write_bytes(excel_bytes)
    if pdf_bytes:
        pdf_path.write_bytes(pdf_bytes)
    metrics = result.get("metrics", {})
    item = {
        "id": stamp,
        "timestamp": metadata.get("timestamp"),
        "usuario": metadata.get("usuario"),
        "sociedad": metadata.get("sociedad"),
        "cuit": metadata.get("cuit"),
        "periodo": metadata.get("periodo"),
        "tipo_auditoria": result.get("tipo_auditoria"),
        "archivo_afip": metadata.get("archivo_afip"),
        "archivo_libro": metadata.get("archivo_libro"),
        "cantidad_archivos_afip": metadata.get("cantidad_archivos_afip"),
        "cantidad_archivos_libro": metadata.get("cantidad_archivos_libro"),
        "excel_path": str(excel_path),
        "pdf_path": str(pdf_path) if pdf_bytes else "",
        "excel_name": excel_name,
        "pdf_name": pdf_name if pdf_bytes else "",
        "excel_b64": _bytes_to_b64(excel_bytes),
        "pdf_b64": _bytes_to_b64(pdf_bytes),
        "metrics": metrics,
        "estado_general": "REVISAR" if metrics.get("observados", 0) or abs(float(metrics.get("diferencia_neta", 0) or 0)) > TOLERANCIA_DEFAULT else "OK",
    }
    hist = load_history()
    hist.insert(0, item)
    save_history(hist)
    return item


def history_label(record: Dict[str, Any], idx: int) -> str:
    periodo = record.get("periodo") or "sin período"
    sociedad = record.get("sociedad") or "sin sociedad"
    tipo = record.get("tipo_auditoria") or "sin tipo"
    fecha = record.get("timestamp") or "sin fecha"
    return f"{idx + 1}. {periodo} · {sociedad} · {tipo} · {fecha}"


def history_to_dataframe(history: List[Dict[str, Any]]) -> pd.DataFrame:
    rows = []
    for r in history:
        m = r.get("metrics", {}) or {}
        rows.append({
            "Período": r.get("periodo", ""),
            "Fecha guardado": r.get("timestamp", ""),
            "Sociedad": r.get("sociedad", ""),
            "CUIT": r.get("cuit", ""),
            "Tipo auditoría": r.get("tipo_auditoria", ""),
            "Estado": r.get("estado_general", ""),
            "IVA AFIP": m.get("iva_afip", m.get("iva_afip_firmado", "")),
            "IVA Libro": m.get("iva_libro", m.get("iva_computado_libro_mes", "")),
            "Dif. neta": m.get("diferencia_neta", ""),
            "Dif. bruta": m.get("diferencia_bruta", ""),
            "OK": m.get("ok", ""),
            "Sólo AFIP": m.get("solo_afip", ""),
            "Sólo Libro": m.get("solo_libro", ""),
            "IVA distinto": m.get("iva_distinto", ""),
            "Archivos AFIP": r.get("cantidad_archivos_afip", 1 if r.get("archivo_afip") else ""),
            "Archivos Libro": r.get("cantidad_archivos_libro", 1 if r.get("archivo_libro") else ""),
            "Archivo AFIP": r.get("archivo_afip", ""),
            "Archivo Libro": r.get("archivo_libro", ""),
        })
    return pd.DataFrame(rows)


def _history_downloads_ui(history: List[Dict[str, Any]]):
    st.markdown("#### Descargar archivos guardados")
    for i, r in enumerate(history):
        label = history_label(r, i)
        excel_bytes = _b64_to_bytes(r.get("excel_b64")) or _read_file_if_exists(r.get("excel_path"))
        pdf_bytes = _b64_to_bytes(r.get("pdf_b64")) or _read_file_if_exists(r.get("pdf_path"))
        excel_name = r.get("excel_name") or (Path(str(r.get("excel_path"))).name if r.get("excel_path") else f"auditoria_{i+1}.xlsx")
        pdf_name = r.get("pdf_name") or (Path(str(r.get("pdf_path"))).name if r.get("pdf_path") else f"auditoria_{i+1}.pdf")
        c0, c1, c2 = st.columns([3.2, 1, 1])
        with c0:
            st.write(label)
        with c1:
            if excel_bytes:
                st.download_button(
                    "⬇️ Excel",
                    data=excel_bytes,
                    file_name=excel_name,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key=f"hist_excel_{i}_{r.get('id','')}",
                    use_container_width=True,
                )
            else:
                st.caption("Excel no disponible")
        with c2:
            if pdf_bytes:
                st.download_button(
                    "⬇️ PDF",
                    data=pdf_bytes,
                    file_name=pdf_name,
                    mime="application/pdf",
                    key=f"hist_pdf_{i}_{r.get('id','')}",
                    use_container_width=True,
                )
            else:
                st.caption("PDF no disponible")


def render_history_panel():
    """Historial principal, estilo app Sueldos: tabla, descargas y administración."""
    with st.expander("Historial guardado", expanded=False):
        hist = load_history()
        if not hist:
            st.info("Todavía no hay auditorías guardadas.")
            return
        hist_df = history_to_dataframe(hist)
        st.dataframe(hist_df, use_container_width=True)
        _history_downloads_ui(hist)
        st.markdown("#### Administrar historial")
        selected_delete = st.multiselect(
            "Seleccioná registros para eliminar",
            options=list(range(len(hist))),
            format_func=lambda i: history_label(hist[i], i),
        )
        c1, c2 = st.columns([1, 1])
        with c1:
            if st.button("Eliminar seleccionados", disabled=not selected_delete):
                delete_set = set(selected_delete)
                remaining = [r for i, r in enumerate(hist) if i not in delete_set]
                save_history(remaining)
                st.success("Registros eliminados.")
                st.rerun()
        with c2:
            confirm_clear = st.checkbox("Confirmo borrar todo el historial")
            if st.button("Borrar todo el historial", disabled=not confirm_clear):
                save_history([])
                st.success("Historial eliminado.")
                st.rerun()

# =============================================================================
# UI
# =============================================================================

def render_metrics(result: Dict[str, Any]):
    m = result["metrics"]

    if result["tipo_auditoria"] == "Auditoría de Comprobantes":
        st.info(
            "Control documental: esta vista responde si cada comprobante existe o no existe en ambos lados. "
            "No valida el IVA mensual por fecha de registración; eso corresponde a la Auditoría IVA del Mes."
        )
        c1, c2, c3, c4 = st.columns(4)
        c1.metric("Claves únicas AFIP", m.get("claves_afip", 0))
        c2.metric("Claves únicas Libro", m.get("claves_libro", 0))
        c3.metric("Comprobantes OK", m.get("ok", 0))
        c4.metric("Comprobantes observados", m.get("observados", 0))

        c5, c6, c7, c8 = st.columns(4)
        c5.metric("Sólo AFIP / ARCA", m.get("solo_afip", 0))
        c6.metric("Sólo Libro IVA", m.get("solo_libro", 0))
        c7.metric("Mismo comprobante, IVA distinto", m.get("iva_distinto", 0))
        c8.metric("Duplicados Libro", m.get("duplicados_libro", 0))

        c9, c10, c11, c12 = st.columns(4)
        c9.metric("IVA en AFIP no cargado", fmt_money(m.get("iva_solo_afip", 0)))
        c10.metric("IVA en Libro sin AFIP", fmt_money(m.get("iva_solo_libro", 0)))
        c11.metric("Dif. IVA en comprobantes encontrados", fmt_money(m.get("iva_diferencia_mismo_comprobante", 0)))
        c12.metric("Diferencia bruta documental", fmt_money(m.get("diferencia_bruta", 0)))
        return

    # Auditoría IVA del Mes
    st.info(
        "Control fiscal del período: esta vista parte del Libro IVA cargado como universo del mes y usa AFIP/ARCA como respaldo documental. "
        "La fecha visible del comprobante se usa para alertar facturas de meses anteriores/posteriores, no para excluirlas automáticamente."
    )
    c1, c2, c3, c4 = st.columns(4)
    c1.metric("IVA computado Libro del mes", fmt_money(m.get("iva_computado_libro_mes", m.get("iva_libro", 0))))
    c2.metric("IVA AFIP respaldado", fmt_money(m.get("iva_afip_encontrado_para_libro", 0)))
    c3.metric("Diferencia neta", fmt_money(m.get("diferencia_neta", 0)))
    c4.metric("Diferencia bruta no compensada", fmt_money(m.get("diferencia_bruta", 0)))

    c5, c6, c7, c8 = st.columns(4)
    c5.metric("IVA Libro no encontrado en AFIP", fmt_money(m.get("iva_libro_no_encontrado_afip", 0)))
    c6.metric("IVA AFIP del mes no registrado", fmt_money(m.get("iva_afip_mes_no_registrado", 0)))
    c7.metric("IVA con fecha fuera período", fmt_money(m.get("iva_fuera_periodo_libro", 0)))
    c8.metric("Comprobantes fuera período", m.get("comprobantes_fuera_periodo_libro", 0))

    c9, c10, c11, c12 = st.columns(4)
    c9.metric("OK", m.get("ok", 0))
    c10.metric("Sólo AFIP", m.get("solo_afip", 0))
    c11.metric("Sólo Libro", m.get("solo_libro", 0))
    c12.metric("IVA distinto", m.get("iva_distinto", 0))

def main():
    st.set_page_config(page_title=APP_TITLE, page_icon="🧾", layout="wide")
    if not login_required():
        return

    st.title(f"{APP_TITLE} · {APP_VERSION}")
    st.caption("Auditoría de comprobantes y validación del IVA mensual")

    with st.sidebar:
        st.header("Sociedad")
        sociedad = st.radio("Elegí una sociedad", SOCIEDADES_DEFAULT, index=2)
        default_cuit = SOCIEDADES_CONFIG.get(sociedad, "")
        cuit_sociedad = st.text_input("CUIT sociedad", value=default_cuit, key=f"cuit_sociedad_{sociedad}")
        st.divider()
        st.caption("Flujo: cargar archivos → elegir auditoría → revisar plan de acción → descargar/guardar historial.")
        st.header("Período fiscal")
        alcance_periodo = st.radio(
            "Alcance",
            ["Mes seleccionado", "Varios meses / consolidado"],
            index=0,
            help=(
                "Mes seleccionado: valida un período fiscal puntual. "
                "Varios meses/consolidado: permite cargar varios archivos por sección y auditar todo junto."
            ),
        )
        today = datetime.today()
        if alcance_periodo == "Mes seleccionado":
            year = st.number_input("Año fiscal", min_value=2020, max_value=2035, value=today.year, step=1)
            month = st.number_input("Mes fiscal", min_value=1, max_value=12, value=today.month, step=1)
            periodo_label = f"{int(month):02d}/{int(year)}"
        else:
            year = today.year
            month = today.month
            periodo_label = "MULTIPERIODO"
            st.caption("Modo consolidado: la app no excluye ni alerta por fecha fuera del mes, porque el universo son todos los archivos cargados.")
        tolerancia = st.number_input("Tolerancia diferencias ($)", min_value=0.0, value=TOLERANCIA_DEFAULT, step=0.01, format="%.2f")

    render_history_panel()

    st.subheader("Paso 1 · Cargar archivos")
    col_a, col_b = st.columns(2)
    with col_a:
        afip_files = st.file_uploader(
            "Archivo/s AFIP / ARCA - Mis Comprobantes",
            type=["xlsx", "xls", "csv"],
            accept_multiple_files=True,
            help="Podés cargar uno o varios archivos AFIP. La app los consolida y conserva el nombre de origen.",
        )
    with col_b:
        libro_files = st.file_uploader(
            "Libro/s IVA del sistema",
            type=["xlsx", "xls", "csv"],
            accept_multiple_files=True,
            help="Podés cargar uno o varios Libros IVA. Útil para auditar varios meses en una sola corrida.",
        )

    if not afip_files or not libro_files:
        st.warning("Cargá al menos un archivo AFIP/ARCA y al menos un Libro IVA para ejecutar la auditoría.")
        return

    try:
        afip_norm = normalize_uploaded_files(afip_files, "AFIP")
        libro_norm = normalize_uploaded_files(libro_files, "LIBRO")
    except Exception as e:
        st.error(f"No se pudo procesar alguno de los archivos: {e}")
        st.stop()

    st.success("Archivos leídos, consolidados y normalizados correctamente.")
    c1, c2, c3, c4, c5, c6 = st.columns(6)
    c1.metric("Archivos AFIP", len(afip_files))
    c2.metric("Archivos Libro", len(libro_files))
    c3.metric("Comprobantes AFIP", len(afip_norm))
    c4.metric("Comprobantes Libro", len(libro_norm))
    c5.metric("IVA AFIP firmado", fmt_money(afip_norm["IVA"].sum()))
    c6.metric("IVA Libro firmado", fmt_money(libro_norm["IVA"].sum()))

    with st.expander("Ver muestra normalizada"):
        st.write("AFIP / ARCA")
        st.caption(file_names_label(afip_files))
        st.dataframe(afip_norm.head(50), use_container_width=True)
        st.write("Libro IVA")
        st.caption(file_names_label(libro_files))
        st.dataframe(libro_norm.head(50), use_container_width=True)

    st.subheader("Paso 2 · Elegir auditoría")
    tipo_seleccionado = st.radio(
        "Seleccioná el informe que querés generar",
        ["Auditoría de Comprobantes", "Auditoría IVA del Mes"],
        horizontal=True,
        help=(
            "Comprobantes = existencia documental por CUIT+Tipo+PV+Número. "
            "IVA del Mes = validación fiscal del Libro IVA del período contra AFIP/ARCA."
        ),
    )

    st.caption(
        "La opción seleccionada queda marcada arriba. El botón rojo de Streamlit ya no se usa como indicador, "
        "para evitar confundir la auditoría activa."
    )

    ejecutar = st.button(f"Ejecutar: {tipo_seleccionado}", type="primary", use_container_width=True)
    if not ejecutar:
        st.info("Elegí una auditoría y presioná Ejecutar. La de comprobantes revisa existencia documental; la de IVA del mes valida el IVA computado en el período del Libro IVA.")
        return

    if tipo_seleccionado == "Auditoría de Comprobantes":
        result = run_auditoria_comprobantes(afip_norm, libro_norm, tolerancia)
    else:
        if alcance_periodo == "Varios meses / consolidado":
            result = run_auditoria_iva_consolidada(afip_norm, libro_norm, tolerancia)
        else:
            result = run_auditoria_iva_mes(afip_norm, libro_norm, int(year), int(month), tolerancia)

    st.subheader(f"Resultado · {result['tipo_auditoria']}")
    render_metrics(result)

    st.subheader("Diagnóstico ejecutivo y plan de acción")
    diag_df = build_diagnostico_ejecutivo(result, metadata={"sociedad": sociedad, "periodo": periodo_label})
    st.dataframe(diag_df, use_container_width=True)
    plan_df = build_plan_accion(result, metadata={"sociedad": sociedad, "periodo": periodo_label})
    st.dataframe(plan_df, use_container_width=True)

    if result["tipo_auditoria"] == "Auditoría IVA del Mes":
        tab0, tab1, tab2, tab3, tab4 = st.tabs(["Control IVA del Mes", "Resumen por estado", "Diferencias por proveedor", "Detalle comprobantes", "Alertas del período"])
        with tab0:
            control_mes = pd.DataFrame([
                ["IVA computado Libro del mes", fmt_money(result["metrics"].get("iva_computado_libro_mes", 0)), "Total del Libro IVA cargado como período fiscal."],
                ["IVA Libro no encontrado en AFIP", fmt_money(result["metrics"].get("iva_libro_no_encontrado_afip", 0)), "Riesgo: crédito fiscal computado sin comprobante encontrado en AFIP/ARCA."],
                ["IVA AFIP del mes no registrado", fmt_money(result["metrics"].get("iva_afip_mes_no_registrado", 0)), "Potencial crédito fiscal omitido en el Libro del mes."],
                ["IVA con fecha fuera del período", fmt_money(result["metrics"].get("iva_fuera_periodo_libro", 0)), "Comprobantes tomados en el período con fecha visible anterior/posterior."],
            ], columns=["Control", "Valor", "Interpretación contable"])
            st.dataframe(control_mes, use_container_width=True)
        with tab1:
            st.dataframe(result["resumen_estado"], use_container_width=True)
        with tab2:
            st.dataframe(result["resumen_proveedor"], use_container_width=True)
        with tab3:
            st.dataframe(result["detalle"], use_container_width=True)
        with tab4:
            st.write("AFIP del mes no registrado en Libro")
            st.dataframe(result.get("afip_mes_no_registrado", pd.DataFrame()), use_container_width=True)
            st.write("Libro con fecha de comprobante fuera del período fiscal seleccionado")
            st.dataframe(result.get("libro_fuera_periodo", pd.DataFrame()), use_container_width=True)
            st.write("Duplicados")
            st.dataframe(result["duplicados"], use_container_width=True)
    else:
        tab1, tab2, tab3, tab4 = st.tabs(["Resumen documental", "Diferencias por proveedor", "Detalle comprobantes", "Duplicados"])
        with tab1:
            st.dataframe(result["resumen_estado"], use_container_width=True)
        with tab2:
            st.dataframe(result["resumen_proveedor"], use_container_width=True)
        with tab3:
            st.dataframe(result["detalle"], use_container_width=True)
        with tab4:
            st.dataframe(result["duplicados"], use_container_width=True)

    metadata = {
        "timestamp": datetime.now().strftime("%d/%m/%Y %H:%M:%S"),
        "usuario": st.session_state.get("auth_user", ""),
        "sociedad": sociedad,
        "cuit": cuit_sociedad,
        "periodo": periodo_label,
        "archivo_afip": file_names_label(afip_files),
        "archivo_libro": file_names_label(libro_files),
        "cantidad_archivos_afip": len(afip_files),
        "cantidad_archivos_libro": len(libro_files),
    }
    excel_bytes = export_excel(result, metadata)
    pdf_bytes = export_pdf(result, metadata)

    st.subheader("Paso 3 · Descargar y guardar")
    safe_periodo = str(periodo_label).replace("/", "_")
    dc1, dc2, dc3 = st.columns(3)
    dc1.download_button(
        "Descargar Excel de Auditoría",
        data=excel_bytes,
        file_name=f"auditoria_iva_{result['tipo_auditoria'].lower().replace(' ', '_')}_{safe_periodo}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
    )
    if pdf_bytes:
        dc2.download_button(
            "Descargar PDF Ejecutivo",
            data=pdf_bytes,
            file_name=f"auditoria_iva_{result['tipo_auditoria'].lower().replace(' ', '_')}_{safe_periodo}.pdf",
            mime="application/pdf",
            use_container_width=True,
        )
    else:
        dc2.warning("PDF no disponible: falta reportlab.")

    if dc3.button("Guardar en Historial", use_container_width=True):
        item = save_run_to_history(result, metadata, excel_bytes, pdf_bytes)
        st.success(f"Auditoría guardada en historial: {item['id']}")
        st.rerun()

    st.caption("Regla central: diferencia neta no reemplaza auditoría. Siempre revisar comprobantes faltantes, duplicados, IVA distinto y diferencia bruta.")


if __name__ == "__main__":
    main()
